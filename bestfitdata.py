import openpyxl
from openpyxl import load_workbook
import exceltojson

"""
Writes year data to the passed workbook

Parameters startYear, endYear: The start and end years (inclusive) to write data
Preconditions: Both startYear and endYear are ints
Parameter workbookName: The name of the workbook to write data to
Precondition: workbookName is a string
Parameter worksheetName: The name of the worksheet in the workbook to write to. Active sheet by default.
Precondition: worksheetName is a string
Parameters row, col: The row and col to start writing at
Preconditions: row and col are both ints
"""
def writeYearData(startYear, endYear, workbookName, worksheetName = 'active', row = 1, col = 1):
    wb = load_workbook(workbookName)
    if worksheetName == 'active':
        ws = wb.active
    else:
        ws = wb[worksheetName]
    for year in range(startYear, endYear+1):
        ws.cell(row = row, column = col, value = year)
        row += 1
    wb.save(workbookName)
    
"""
Returns an array containing the compound score of each year

Parameter json: The JSON object containing Billboard Sentiment
Precondition: json is a valid Billboard Sentiment JSON object
"""
def getCompoundScoreArray(json):
    compoundScoreArr = []
    for year in json:
        yearlyCompoundScore = 0
        count = 0
        for block in json[year]:
            for song in block:
                yearlyCompoundScore += float(block[song]['score'])
                count += 1
        compoundScoreArr.append(yearlyCompoundScore / count)
    return compoundScoreArr
                

"""
Writes yearly VADER compound score data to the passed workbook
Uses active sheet, starts at row 1, column 2 by default
"""
def writeCompoundScoreData(compoundScoreArr, workbookName, worksheetName = 'active', row = 1, col = 2):
    wb = load_workbook(workbookName)
    if worksheetName == 'active':
        ws = wb.active
    else:
        ws = wb[worksheetName]
    for score in compoundScoreArr:
        ws.cell(row = row, column = col, value = score)
        row += 1
    wb.save(workbookName)
    