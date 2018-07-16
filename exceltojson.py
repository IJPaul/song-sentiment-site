from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os

"""
Writes a JSON object to a .json file

Parameter fileName: The name of the file to write data to
Precondition: fileName is a string with no extension
Parameter jsonData: The JSON object to write
Precondition: jsonData is a valid JSON object
"""
def writeJsonToFile(fileName, jsonData):
    filepath = os.getcwd() + '/' + fileName + '.json'
    with open(filepath, 'w') as fp:
        json.dump(jsonData, fp)

"""
Converts values in an excel workbook to a Billboard Sentiment JSON object. This function
is specific to the excel workbook with Billbaord Data for this project

Parameter workbookName: The name of the excel workbook file
Precondition: workbookName is a string
Parameters startRow, endRow: The start and end row to read data from
Preconditions: Both startRow and endRow are ints
Parameters startCol, endCol: The start and end columns to read data from
Preconditions: Both startCol and endCol are strings
"""
def xlsxToJson(workbookName):
    wb = load_workbook(workbookName)
    sheetsArr = wb.sheetnames
    JsonObj = {}
    for sheetName in sheetsArr:
        rowNum = 1
        year = sheetName.split('.xlsx')[0]
        JsonObj[year] = []
        ws = wb[sheetName]
        for row in ws.rows:
            billboardEntry = {}
            billboardEntry[str(rowNum)] = {}
            entryKeys = ['artist', 'song', 'score', 'mood']
            colNum = 1
            for cell in row:
                key = entryKeys[colNum-1]
                billboardEntry[str(rowNum)][key] = cell.value
                colNum += 1
            JsonObj[year].append(billboardEntry)
            rowNum += 1
    return JsonObj