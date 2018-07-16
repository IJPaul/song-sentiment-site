import sentiment
import scrapebillboard
from openpyxl import Workbook
from openpyxl import load_workbook


WORKBOOK_NAME = 'BillboardSentiment.xlsx'
wb = load_workbook(WORKBOOK_NAME)
wb.save(WORKBOOK_NAME)

"""
Writes the top 100 songs to an excel sheet along with the artist, compound score of the song,
and overall mood of the song (negative, positive, or neutral), between the specified years (inclusive)

Parameters startYear, endYear: the first year to start collecting data on, the last year to collect data on
Preconditons: both startYear and endYear are ints 1940 <= startYear, endYear <= 2017
"""
def writeBillboardData(startYear, endYear):
    for year in range(startYear, endYear+1):
        row = 1
        
        sheetName = str(year) + '.xlsx'
        if not sheetName in wb.sheetnames:
            ws = wb.create_sheet(sheetName)
        else:
            ws = wb[sheetName]
        
        if year == 2013 or year == 2017:
            topArtistAndSongs = scrapebillboard.getArtistsAndSongs(year)
        else:
            billboardUrl = sentiment.getBillboardUrl(year)
            topArtistAndSongs = sentiment.getTopSongs(billboardUrl)
        for i in range(0, len(topArtistAndSongs)):
            
            artistAndSong = topArtistAndSongs[i]
            artist = artistAndSong[0]
            song = artistAndSong[1]
            
            geniusUrl = sentiment.getGeniusUrl(artist, song)
            if not geniusUrl is None:
        
                songSentiment = sentiment.getSongSentiment(geniusUrl)
                compoundScore = songSentiment['compound']
                mood = sentiment.getSongMood(geniusUrl)
                
                arr = [artist, song, compoundScore, mood]
                
                for col in range(1, len(arr)+1):
                    ws.cell(row = row, column = col, value = arr[col-1])
                    wb.save(WORKBOOK_NAME)
                row += 1
        print('done' + str(year))
    